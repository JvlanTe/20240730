import openpyxl

wb = openpyxl.load_workbook("files/new_wb.xlsx")

ws = wb.active

ws["A1"].value = 5
ws["B1"].value = 10
ws["C1"].value = "=A1*B1"

print(ws["C1"].value)

wb.save("files/new_wb2.xlsx")
